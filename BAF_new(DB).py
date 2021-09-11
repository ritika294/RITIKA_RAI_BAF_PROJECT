
import pyomo
import cx_Oracle
import pandas as pd
import numpy as np
import pyomo.opt
import pyomo.environ as pe
import os
import math
from pyomo.environ import *
from openpyxl import load_workbook
import sys

grade_val = sys.argv[1]

print("running optimizer for grade = ", grade_val)

def readdata(conn,querystring):
    return pd.read_sql(querystring, con=conn) 


os.chdir("C:/Users/ritika .rai/Documents/Pyomo CBC Solver")
#data="C:/Users/ritika.rai/Documents/Inventory.xlsx"
conn = cx_Oracle.connect('crmqry/crmqry@133.0.1.2/crm2')
shipquery = "select ccl_id_coil , ccl_cd_status,ccl_sec2,ccl_sec1,ccl_ms_piece_actl from v_cold_coil where ccl_cd_status in ('BC','BB','PP','BF','AP')"
shipment=readdata(conn,shipquery)


details1=pd.read_excel(data, sheet_name='Sheet1')
print(details1)

m1=ConcreteModel()
m1.cn=Set(initialize=details1['Id Coil'])
m1.levels=Set(initialize=[i for i in range(1,6)])
a=[]
nw1={}
nw2={}
nw3={}
details1.reset_index(inplace=True)
if 'index' in details1.columns:
    details1.drop('index',axis=1,inplace=True)
c=0
for i in m1.cn:
#     print('--------------',i)
#     print('>>>>>>>>>>>',c)
    nw1.update({i:details1['Mass'][c]})
    nw2.update({i:details1['Width'][c]})
    nw3.update({i:details1['OD'][c]})
    c=c+1
m1.p_mass=Param(m1.cn,initialize=nw1)
m1.p_width=Param(m1.cn,initialize=nw2)
m1.p_OD=Param(m1.cn,initialize=nw3)

m1.V_binVar=Var(m1.cn,m1.levels,domain=Binary)

def horizontal(m1,cn):
    return sum(m1.V_binVar[cn,lev] for lev in m1.levels)<=1
m1.c_horizontalSum=Constraint(m1.cn,rule=horizontal)

def vertical(m1,lev):
    return sum(m1.V_binVar[cn,lev] for cn in m1.cn)<=1
m1.c_verticalSum=Constraint(m1.levels,rule=vertical)

def sumOfWidth(m1):
    return sum(m1.p_width[cn]*m1.V_binVar[cn,lev] for cn in m1.cn for lev in m1.levels)<=6000
m1.c_widthSum=Constraint(rule=sumOfWidth)

def masssum(m1):
    return sum(m1.V_binVar[cn,lev]*m1.p_mass[cn] for cn in m1.cn for lev in m1.levels)<=107
m1.c_massCon=Constraint(rule=masssum)

def objectiverule(m1):
    return sum(m1.V_binVar[cn,lev]*m1.p_mass[cn] for cn in m1.cn for lev in m1.levels)
m1.objective=Objective(rule=objectiverule, sense=maximize)


def pyomo_postprocess(options=None, instance=None, results=None): 
    print('CN',"           ",'Level',"   ",'Allowed',"   ",'Width',"    ",'Mass',"       ",'OD')  #cn >>> Coil no.
    CN=[]
    LEV=[]
    BINV=[]
    WIDTH=[]
    MASS=[]
    OD=[]
    for cn in m1.cn:         #cn >>> Coil no.
        for lev in m1.levels:
            if(m1.V_binVar[cn,lev].value>0):
                a.append(cn)
                CN.append(cn);LEV.append(lev);BINV.append(m1.V_binVar[cn,lev].value);WIDTH.append(m1.p_width[cn]);MASS.append(m1.p_mass[cn]);OD.append(m1.p_OD[cn])
                print(cn,"     ",lev,"       ",m1.V_binVar[cn,lev].value,"     ",m1.p_width[cn],"     ",m1.p_mass[cn],"     ",m1.p_OD[cn])
                pass
            pass
        pass
    frame={'CN':CN,
           'Level':LEV,
           'Allowed':BINV,
           'Width':WIDTH,
           'Mass':MASS,
           'OD':OD}
    df = pd.DataFrame(frame)
    return df

                

from pyomo.opt import SolverFactory
import pyomo.environ
opt = SolverFactory("cbc")
opt.options['sec']=600
opt.options['ratio']=0.05
results = opt.solve(m1,tee=True)
results.write()
print("\nDisplaying Solution\n" + '-'*60)
post_proc_df=pyomo_postprocess(None, m1, results)

def Cluster(x):
    if 1070 < x <= 1091:
        return 1
    elif 1091 < x <= 1111:
        return 2
    elif 1111 < x <= 1131:
        return 3
    elif 1131 < x <= 1151:
        return 4
    elif 1151 < x <= 1171:
        return 5
    elif 1171 < x <= 1191:
        return 6
    elif 1191 < x <= 1211:
        return 7
    elif 1211 < x <= 1231:
        return 8
    elif 1231 < x <= 1251:
        return 9
    elif 1251 < x <= 1271:
        return 10
    elif 1271 < x <= 1291:
        return 11
    elif 1291 < x <= 1311:
        return 12
    elif 1311 < x <= 1331:
        return 13
    elif 1331 < x <= 1351:
        return 14
    elif 1351 < x <= 1371:
        return 15
    elif 1371 < x <= 1391:
        return 16
    elif 1391 < x <= 1411:
        return 17
    elif 1411 < x <= 1431:
        return 18
    elif 1431 < x <= 1451:
        return 19
    elif 1451 < x <= 1471:
        return 20
    elif 1471 < x <= 1491:
        return 21
    elif 1491 < x <= 1511:
        return 22
    elif 1511 < x <= 1531:
        return 23
    elif 1551 < x <= 1571:
        return 24
    elif 1571 < x <= 1591:
        return 25
    elif 1591 < x <= 1611:
        return 26
    elif 1611 < x <= 1631:
        return 27
    elif 1631 < x <= 1651:
        return 28
    elif 1651 < x <= 1671:
        return 29
    elif 1671 < x <= 1691:
        return 30
    elif 1691 < x <= 1711:
        return 31
    elif 1711 < x <= 1731:
        return 32
    elif 1731 < x <= 1751:
        return 33
    elif 1751 < x <= 1771:
        return 34
    elif 1771 < x <= 1791:
        return 35
    elif 1791 < x <= 1811:
        return 36
    elif 1811 < x <= 1831:
        return 37
    elif 1831 < x <= 1851:
        return 38
    elif 1851 < x <= 1871:
        return 39
    elif 1871 < x <= 1891:
        return 40
    elif 1891 < x <= 1911:
        return 41

    else:
        pass

post_proc_df['Cluster'] = post_proc_df.apply(lambda x: Cluster(x['OD']), axis = 1)

Sort_post_proc = post_proc_df.sort_values(by=['Cluster','Mass'])
print(Sort_post_proc)

#if input().lower()=='yes':
      
for cn in a:
        details1=details1[details1['Id Coil'] != cn]
    #     print(details1.tail())
writer = pd.ExcelWriter('C:/Users/shikhar.mishra/Documents/Inventory.xlsx')
details1.to_excel(writer)
writer.save()
writer1 = pd.ExcelWriter('C:/Users/shikhar.mishra/Documents/BAFout17.xlsx')
Sort_post_proc.to_excel(writer1)
writer1.save()


#{25:details1['Mass'][0]}

file_path = 'C:/Users/shikhar.mishra/Documents/BAFout.xlsx'


def append_df_to_excel(filename, df, sheet_name='BAF_OUT', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
#    from openpyxl import load_workbook
#
#    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

#        # truncate sheet
#        if truncate_sheet and sheet_name in writer.book.sheetnames:
#            # index of [sheet_name] sheet
#            idx = writer.book.sheetnames.index(sheet_name)
#            # remove [sheet_name]
#            writer.book.remove(writer.book.worksheets[idx])
#            # create an empty sheet [sheet_name] using old index
#            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    print('Done----->')


# calling function
append_df_to_excel(filename=file_path, df=Sort_post_proc, sheet_name='BAF_OUT', startrow=1,
                       truncate_sheet=False,index=False)
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
dfList = post_proc_df
path = 'C:/Users/ritika.rai/Documents/Inventory.xlsx'
newpath = 'C:/Users/ritika.rai/Documents/Inventory.xlsx'

#for fn in os.listdir(path): 
#  # Absolute file path
#  file = os.path.join(path, fn)
#  if os.path.isfile(file): 
#    # Import the excel file and call it xlsx_file 
#    xlsx_file = pd.ExcelFile(file) 
#    # View the excel files sheet names 
#    xlsx_file.sheet_names 
#    # Load the xlsx files Data sheet as a dataframe 
#    df = xlsx_file.parse('Sheet1',header= None) 
#    df_NoHeader = df[2:] 
#    data = df_NoHeader 
#    # Save individual dataframe
#    data.to_excel(os.path.join(newpath, fn))
#
#    dfList.append(data) 
#
#appended_data = pd.concat(dfList)
#
#appended_data.to_excel(os.path.join(newpath, 'C:/Users/shikhar.mishra/Documents/master_data.xlsx'))

